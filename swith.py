´"""Reconcile a purchase master workbook against supplier statements.

Dependencies: openpyxl, pdfplumber, and LibreOffice (only for legacy .xls files).

Example:
    python3 reconcile_statements.py \
        --main purchase_master.xlsx \
        --statements supplier_statements \
        --output reconciliation_result.xlsx \
        --month 2026-6
"""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import subprocess
import tempfile
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
MAIN_PATTERN = "采购明细表*.xlsx"
DEFAULT_STATEMENT_DIR = ROOT / "6月份对账单" / "已OK"
DEFAULT_DATA_JSON = ROOT / "outputs" / "reconciliation_data.json"
DEFAULT_RESULT_XLSX = ROOT / "6月份供应商对账核对结果.xlsx"

QTY_TOLERANCE = 0.000001
AMOUNT_TOLERANCE = 0.01


SUPPLIER_ALIASES = {
    "亿隆/跃鸿": ["亿隆", "跃鸿", "亿隆/跃鸿"],
    "酷影": ["酷影", "酷影视界"],
}


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return re.sub(r"\s+", "", text).strip()


def norm_code(value: Any) -> str:
    text = norm_text(value)
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.upper()


def to_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return 0.0
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = (
        text.replace(",", "")
        .replace("￥", "")
        .replace("¥", "")
        .replace("RMB", "")
        .replace(" ", "")
    )
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        return float(match.group(0)) if match else 0.0


def round2(value: float) -> float:
    result = round(float(value or 0), 2)
    return 0.0 if result == 0 else result


def round6(value: float) -> float:
    result = round(float(value or 0), 6)
    return 0.0 if result == 0 else result


@dataclass
class Agg:
    supplier: str = ""
    rows: int = 0
    qty: float = 0.0
    amount: float = 0.0
    codes: set[str] = field(default_factory=set)
    by_code: dict[str, dict[str, Any]] = field(default_factory=dict)

    def add(self, code: str, name: str, qty: float, amount: float) -> None:
        code = norm_code(code)
        name = norm_text(name)
        self.rows += 1
        self.qty += qty
        self.amount += amount
        if code:
            self.codes.add(code)
        bucket = self.by_code.setdefault(
            code or "(空编码)",
            {"code": code or "(空编码)", "name": name, "qty": 0.0, "amount": 0.0, "rows": 0},
        )
        if name and not bucket.get("name"):
            bucket["name"] = name
        bucket["qty"] += qty
        bucket["amount"] += amount
        bucket["rows"] += 1


@dataclass
class Statement:
    path: Path
    kind: str
    supplier_text: str = ""
    parse_status: str = "ok"
    parse_note: str = ""
    agg: Agg = field(default_factory=Agg)
    header_row: int | None = None
    sheet_name: str = ""


def find_default_main(root: Path = ROOT) -> Path:
    main_files = [path for path in root.glob(MAIN_PATTERN) if not path.name.startswith("~$")]
    if len(main_files) != 1:
        raise RuntimeError(f"当前目录应只有一个 {MAIN_PATTERN}，实际找到 {len(main_files)} 个")
    return main_files[0]


def read_main(
    main_path: Path,
    header_row: int = 3,
    data_row: int = 4,
    sheet_name: str | None = None,
) -> dict[str, Agg]:
    wb = load_workbook(main_path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"总表不存在工作表：{sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    headers = [norm_text(c.value) for c in ws[header_row]]
    required = ["供应商", "商品编码", "商品名称", "数量", "价税合计"]
    positions = {name: headers.index(name) for name in required if name in headers}
    missing = [name for name in required if name not in positions]
    if missing:
        raise RuntimeError(f"总表缺少字段：{', '.join(missing)}")

    result: dict[str, Agg] = {}
    for row in ws.iter_rows(min_row=data_row, values_only=True):
        supplier = norm_text(row[positions["供应商"]])
        if not supplier:
            continue
        agg = result.setdefault(supplier, Agg(supplier=supplier))
        agg.add(
            row[positions["商品编码"]],
            row[positions["商品名称"]],
            to_number(row[positions["数量"]]),
            to_number(row[positions["价税合计"]]),
        )
    wb.close()
    return result


def row_values(ws, row_idx: int) -> list[Any]:
    try:
        return [c.value for c in ws[row_idx]]
    except IndexError:
        return []


def safe_cell(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def header_positions(values: list[Any]) -> dict[str, int]:
    normalized = [norm_text(v) for v in values]
    aliases = {
        "code": ["物料料号", "商品编码", "产品编号", "物料编码", "料号"],
        "name": ["物料名称", "商品名称", "产品名称", "品名"],
        "qty": ["数量", "入库数量"],
        "amount": ["金额", "价税合计", "含税金额", "总金额"],
    }
    pos: dict[str, int] = {}
    for key, names in aliases.items():
        for name in names:
            if name in normalized:
                pos[key] = normalized.index(name)
                break
    return pos


def find_header(ws) -> tuple[int | None, dict[str, int]]:
    for idx in range(1, min(ws.max_row, 30) + 1):
        pos = header_positions(row_values(ws, idx))
        if "qty" in pos and "amount" in pos and ("code" in pos or "name" in pos):
            return idx, pos
    return None, {}


def period_score(text: str, month: str | None) -> int:
    if not month:
        return 0
    numbers = [int(value) for value in re.findall(r"\d+", month)]
    if not numbers:
        return 0
    year = numbers[0] if numbers[0] >= 1000 else None
    month_number = numbers[1] if year and len(numbers) > 1 else numbers[0]
    month_pattern = rf"0?{month_number}"
    if year:
        pattern = rf"{year}\s*[-_./年]?\s*{month_pattern}\s*(?:月|$)"
        return 200 if re.search(pattern, text, re.IGNORECASE) else 0
    return 80 if re.search(rf"(^|\D){month_pattern}\s*月", text) else 0


def sheet_score(ws, header_row: int, month: str | None = None) -> int:
    title = norm_text(ws.title).lower()
    score = 200 if getattr(ws, "sheet_state", "visible") == "visible" else 0
    score += period_score(title, month)
    if "格式" in title or "模板" in title:
        score -= 300
    top_text = ""
    for idx in range(1, min(header_row, 5) + 1):
        top_text += "".join(norm_text(v) for v in row_values(ws, idx) if v is not None)
    score += period_score(top_text, month) // 2
    return score


def convert_xls(path: Path, tmp_dir: Path) -> Path:
    soffice = shutil.which("soffice")
    if not soffice:
        mac_soffice = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
        if mac_soffice.exists():
            soffice = str(mac_soffice)
    if not soffice:
        raise RuntimeError("读取 .xls 需要安装 LibreOffice，且系统中未找到 soffice")
    subprocess.run(
        [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(tmp_dir), str(path)],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    converted = tmp_dir / f"{path.stem}.xlsx"
    if not converted.exists():
        candidates = list(tmp_dir.glob("*.xlsx"))
        if not candidates:
            raise RuntimeError("xls conversion did not produce xlsx")
        converted = candidates[0]
    return converted


def parse_excel_statement(
    path: Path,
    tmp_dir: Path | None = None,
    month: str | None = None,
) -> Statement:
    actual = path
    statement = Statement(path=path, kind=path.suffix.lower().lstrip("."))
    try:
        if path.suffix.lower() == ".xls":
            if tmp_dir is None:
                raise RuntimeError("tmp_dir required for xls")
            actual = convert_xls(path, tmp_dir)
        wb = load_workbook(actual, data_only=True, read_only=True)
        best_sheet = None
        best_header = None
        best_pos = None
        best_score = -10_000
        for ws in wb.worksheets:
            header, pos = find_header(ws)
            if header:
                current_score = sheet_score(ws, header, month)
                if current_score > best_score:
                    best_sheet, best_header, best_pos = ws, header, pos
                    best_score = current_score
        if best_sheet is None or best_header is None or best_pos is None:
            statement.parse_status = "partial"
            statement.parse_note = "未识别到数量/金额表头"
            wb.close()
            return statement
        statement.header_row = best_header
        statement.sheet_name = best_sheet.title
        supplier_area: list[str] = []
        for row_index in range(1, min(best_header, 10) + 1):
            supplier_area.extend(
                norm_text(value)
                for value in row_values(best_sheet, row_index)[:20]
                if norm_text(value)
            )
        statement.supplier_text = "|".join(supplier_area)
        agg = Agg()
        for row in best_sheet.iter_rows(min_row=best_header + 1, values_only=True):
            code = safe_cell(row, best_pos.get("code"))
            name = safe_cell(row, best_pos.get("name"))
            qty = to_number(safe_cell(row, best_pos.get("qty")))
            amount = to_number(safe_cell(row, best_pos.get("amount")))
            if not code and not name:
                continue
            if norm_text(code) in {"物料料号", "商品编码"}:
                continue
            agg.add(code, name, qty, amount)
        statement.agg = agg
        if agg.rows == 0:
            statement.parse_status = "partial"
            statement.parse_note = "识别到表头但未提取到有效明细"
        wb.close()
        return statement
    except Exception as exc:
        statement.parse_status = "error"
        statement.parse_note = f"{type(exc).__name__}: {exc}"
        return statement


def parse_pdf_statement(path: Path) -> Statement:
    statement = Statement(path=path, kind="pdf")
    try:
        text_parts: list[str] = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text_parts.append(page.extract_text() or "")
        text = "\n".join(text_parts)
        first_line = next((norm_text(line) for line in text.splitlines() if norm_text(line)), "")
        statement.supplier_text = first_line
        agg = Agg()
        for line in text.splitlines():
            stripped = line.strip()
            if not re.match(r"^\d+\s+", stripped):
                continue
            if "CGDD" not in stripped and "CG" not in stripped:
                continue
            tokens = stripped.split()
            code = ""
            for token in tokens:
                if re.match(r"^[A-Za-z]{1,4}\d", token):
                    code = token
                    break
            if "PCS" in tokens:
                pcs_idx = tokens.index("PCS")
                numeric_after_pcs = []
                for token in tokens[pcs_idx + 1 :]:
                    cleaned = token.replace("￥", "").replace(",", "")
                    if re.fullmatch(r"-?\d+(?:\.\d+)?", cleaned):
                        numeric_after_pcs.append(cleaned)
                    if len(numeric_after_pcs) >= 3:
                        break
                if len(numeric_after_pcs) < 3:
                    continue
                qty = to_number(numeric_after_pcs[0])
                amount = to_number(numeric_after_pcs[2])
            else:
                nums = re.findall(r"-?\d[\d,]*(?:\.\d+)?", stripped.replace("￥", ""))
                if len(nums) < 3:
                    continue
                qty = to_number(nums[-3])
                amount = to_number(nums[-1])
            if not code:
                continue
            name = ""
            if code:
                after_code = stripped.split(code, 1)[-1]
                name = after_code.split("PCS", 1)[0].strip()
            agg.add(code, name, qty, amount)
        statement.agg = agg
        if agg.rows == 0:
            statement.parse_status = "partial"
            statement.parse_note = "PDF 未提取到有效明细"
        return statement
    except Exception as exc:
        statement.parse_status = "error"
        statement.parse_note = f"{type(exc).__name__}: {exc}"
        return statement


def parse_statements(statement_dir: Path, month: str | None = None) -> list[Statement]:
    if not statement_dir.is_dir():
        raise RuntimeError(f"子表目录不存在：{statement_dir}")
    statements: list[Statement] = []
    with tempfile.TemporaryDirectory(prefix="statement_xls_") as tmp:
        tmp_dir = Path(tmp)
        paths = [
            path
            for path in statement_dir.rglob("*")
            if path.is_file() and not path.name.startswith((".", "~$"))
        ]
        for path in sorted(paths, key=lambda item: str(item).lower()):
            if path.suffix.lower() in {".xlsx", ".xls"}:
                statements.append(parse_excel_statement(path, tmp_dir, month))
            elif path.suffix.lower() == ".pdf":
                statements.append(parse_pdf_statement(path))
            else:
                statements.append(
                    Statement(path=path, kind=path.suffix.lower().lstrip("."), parse_status="error", parse_note="不支持的文件类型")
                )
    return statements


def load_aliases(path: Path | None) -> dict[str, list[str]]:
    aliases = {key: list(values) for key, values in SUPPLIER_ALIASES.items()}
    if not path:
        return aliases
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise RuntimeError("供应商别名 JSON 必须是对象，例如 {\"总表供应商\": [\"文件别名\"]}")
    for supplier, values in payload.items():
        if isinstance(values, str):
            values = [values]
        if not isinstance(values, list):
            raise RuntimeError(f"供应商 {supplier} 的别名必须是字符串或字符串数组")
        aliases[str(supplier)] = [str(value) for value in values]
    return aliases


def aliases_for(supplier: str, aliases: dict[str, list[str]]) -> list[str]:
    return [supplier] + aliases.get(supplier, [])


def name_score(supplier: str, statement: Statement, aliases: dict[str, list[str]]) -> int:
    haystacks = [norm_text(statement.path.stem), norm_text(statement.supplier_text)]
    score = 0
    for alias in aliases_for(supplier, aliases):
        alias_norm = norm_text(alias)
        if not alias_norm:
            continue
        if any(alias_norm in hay for hay in haystacks):
            score += 100
        if "/" in alias_norm:
            parts = [p for p in alias_norm.split("/") if p]
            if parts and all(any(part in hay for hay in haystacks) for part in parts):
                score += 120
    return score


def statement_score(
    main: Agg,
    statement: Statement,
    aliases: dict[str, list[str]],
    qty_tolerance: float,
    amount_tolerance: float,
) -> tuple[float, str]:
    score = float(name_score(main.supplier, statement, aliases))
    reasons: list[str] = []
    if score:
        reasons.append("名称匹配")
    code_overlap = len(main.codes & statement.agg.codes)
    if code_overlap:
        score += min(80, code_overlap * 8)
        reasons.append(f"编码重叠{code_overlap}个")
    qty_diff = abs(main.qty - statement.agg.qty)
    amt_diff = abs(main.amount - statement.agg.amount)
    if qty_diff <= qty_tolerance:
        score += 30
        reasons.append("数量一致")
    if amt_diff <= amount_tolerance:
        score += 30
        reasons.append("金额一致")
    return score, "，".join(reasons)


def match_statements(
    main_aggs: dict[str, Agg],
    statements: list[Statement],
    aliases: dict[str, list[str]],
    qty_tolerance: float,
    amount_tolerance: float,
) -> tuple[dict[str, tuple[Statement, str] | None], set[Path]]:
    matches: dict[str, tuple[Statement, str] | None] = {}
    used: set[Path] = set()
    scored: list[tuple[float, str, str, Statement]] = []
    for supplier, main in main_aggs.items():
        for statement in statements:
            score, reason = statement_score(main, statement, aliases, qty_tolerance, amount_tolerance)
            if score >= 60:
                scored.append((score, supplier, reason, statement))
    for score, supplier, reason, statement in sorted(scored, key=lambda item: item[0], reverse=True):
        if supplier in matches or statement.path in used:
            continue
        matches[supplier] = (statement, reason)
        used.add(statement.path)
    for supplier in main_aggs:
        matches.setdefault(supplier, None)
    return matches, used


def make_detail_rows(
    main_aggs: dict[str, Agg],
    matches: dict[str, tuple[Statement, str] | None],
    qty_tolerance: float,
    amount_tolerance: float,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for supplier in sorted(main_aggs):
        match = matches[supplier]
        if not match:
            main = main_aggs[supplier]
            for code, item in sorted(main.by_code.items()):
                rows.append([supplier, code, item.get("name", ""), round(item["qty"], 6), 0, round(item["qty"], 6), round2(item["amount"]), 0, round2(item["amount"]), "子表缺失"])
            continue
        statement, _reason = match
        main = main_aggs[supplier]
        all_codes = sorted(set(main.by_code) | set(statement.agg.by_code))
        for code in all_codes:
            m = main.by_code.get(code, {"qty": 0.0, "amount": 0.0, "name": ""})
            s = statement.agg.by_code.get(code, {"qty": 0.0, "amount": 0.0, "name": ""})
            qty_diff = m["qty"] - s["qty"]
            amt_diff = m["amount"] - s["amount"]
            if abs(qty_diff) <= qty_tolerance and abs(amt_diff) <= amount_tolerance:
                continue
            if code not in main.by_code:
                diff_type = "总表缺少此编码"
            elif code not in statement.agg.by_code:
                diff_type = "子表缺少此编码"
            else:
                diff_type = "数量/金额不一致"
            rows.append([
                supplier,
                code,
                m.get("name") or s.get("name") or "",
                round(m["qty"], 6),
                round(s["qty"], 6),
                round(qty_diff, 6),
                round2(m["amount"]),
                round2(s["amount"]),
                round2(amt_diff),
                diff_type,
            ])
    return rows


def relative_name(path: Path, base: Path) -> str:
    try:
        return str(path.relative_to(base))
    except ValueError:
        return path.name


def build_output(
    main_path: Path,
    statement_dir: Path,
    output_xlsx: Path,
    month: str | None = None,
    main_sheet: str | None = None,
    main_header_row: int = 3,
    main_data_row: int = 4,
    qty_tolerance: float = QTY_TOLERANCE,
    amount_tolerance: float = AMOUNT_TOLERANCE,
    aliases: dict[str, list[str]] | None = None,
) -> dict[str, Any]:
    aliases = aliases or load_aliases(None)
    main_aggs = read_main(main_path, main_header_row, main_data_row, main_sheet)
    statements = parse_statements(statement_dir, month)
    matches, used = match_statements(
        main_aggs,
        statements,
        aliases,
        qty_tolerance,
        amount_tolerance,
    )
    summary: list[list[Any]] = []
    for supplier in sorted(main_aggs):
        main = main_aggs[supplier]
        match = matches[supplier]
        if not match:
            summary.append([supplier, round6(main.qty), None, None, round2(main.amount), None, None, "", "未完成-未找到对账单", "未找到可匹配的供应商对账单"])
            continue
        statement, reason = match
        qty_diff = main.qty - statement.agg.qty
        amt_diff = main.amount - statement.agg.amount
        status = "完成" if abs(qty_diff) <= qty_tolerance and abs(amt_diff) <= amount_tolerance else "未完成-金额/数量不一致"
        notes = [reason]
        if statement.sheet_name:
            notes.append(f"读取工作表：{statement.sheet_name}")
        if statement.parse_status != "ok":
            status = "需复核"
            notes.append(statement.parse_note)
        if status == "完成" and main.codes != statement.agg.codes:
            notes.append("存在编码差异")
        summary.append([
            supplier,
            round6(main.qty),
            round6(statement.agg.qty),
            round6(qty_diff),
            round2(main.amount),
            round2(statement.agg.amount),
            round2(amt_diff),
            relative_name(statement.path, statement_dir),
            status,
            "；".join(filter(None, notes)),
        ])

    detail = make_detail_rows(main_aggs, matches, qty_tolerance, amount_tolerance)
    unmatched_files: list[list[Any]] = []
    for statement in sorted(statements, key=lambda s: s.path.name):
        if statement.path in used:
            continue
        possible = []
        for supplier, main in main_aggs.items():
            score, reason = statement_score(main, statement, aliases, qty_tolerance, amount_tolerance)
            if score > 0:
                possible.append((score, supplier, reason))
        possible_text = "；".join(f"{supplier}({reason})" for _score, supplier, reason in sorted(possible, reverse=True)[:3])
        reason = statement.parse_note if statement.parse_status != "ok" else "未被任何总表供应商采用"
        unmatched_files.append([relative_name(statement.path, statement_dir), reason, possible_text])

    result = {
        "summary_headers": ["供应商", "总表数量", "子表数量", "数量差异", "总表金额", "子表金额", "金额差异", "匹配文件", "状态", "备注"],
        "summary": summary,
        "detail_headers": ["供应商", "商品编码/物料料号", "商品名称", "总表数量", "子表数量", "数量差异", "总表金额", "子表金额", "金额差异", "差异类型"],
        "detail": detail,
        "unmatched_headers": ["文件名", "原因", "可能对应供应商"],
        "unmatched": unmatched_files,
        "meta": {
            "main_supplier_count": len(main_aggs),
            "statement_file_count": len(statements),
            "xlsx_count": sum(1 for s in statements if s.path.suffix.lower() == ".xlsx"),
            "xls_count": sum(1 for s in statements if s.path.suffix.lower() == ".xls"),
            "pdf_count": sum(1 for s in statements if s.path.suffix.lower() == ".pdf"),
            "unsupported_count": sum(1 for s in statements if s.path.suffix.lower() not in {".xlsx", ".xls", ".pdf"}),
            "main_path": str(main_path),
            "statement_dir": str(statement_dir),
            "result_xlsx": str(output_xlsx),
            "month": month or "",
            "qty_tolerance": qty_tolerance,
            "amount_tolerance": amount_tolerance,
        },
    }
    return result


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Hiragino Sans GB", size=10, color="FFFFFF", bold=True)
BODY_FONT = Font(name="Hiragino Sans GB", size=10)
STATUS_FILLS = {
    "完成": PatternFill("solid", fgColor="E2F0D9"),
    "未完成-金额/数量不一致": PatternFill("solid", fgColor="FCE4D6"),
    "未完成-未找到对账单": PatternFill("solid", fgColor="FFF2CC"),
    "需复核": PatternFill("solid", fgColor="DDEBF7"),
}
DIFF_FILL = PatternFill("solid", fgColor="FCE4D6")


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        ws.auto_filter.ref = ws.dimensions
        return
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_data_sheet(
    ws,
    widths: list[float],
    qty_columns: set[int] | None = None,
    amount_columns: set[int] | None = None,
) -> None:
    qty_columns = qty_columns or set()
    amount_columns = amount_columns or set()
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 26
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(
                horizontal="right" if cell.column in qty_columns | amount_columns else "left",
                vertical="top",
                wrap_text=cell.column not in qty_columns | amount_columns,
            )
            if cell.column in qty_columns:
                cell.number_format = "#,##0.######"
            elif cell.column in amount_columns:
                cell.number_format = "#,##0.00"


def write_workbook(data: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "汇总结果"
    summary_ws.append(data["summary_headers"])
    for row in data["summary"]:
        summary_ws.append(row)
    style_data_sheet(
        summary_ws,
        [18, 14, 14, 14, 16, 16, 16, 42, 27, 55],
        qty_columns={2, 3, 4},
        amount_columns={5, 6, 7},
    )
    add_table(summary_ws, "ReconciliationSummary")
    for row_index in range(2, summary_ws.max_row + 1):
        status = summary_ws.cell(row_index, 9).value
        if status in STATUS_FILLS:
            summary_ws.cell(row_index, 9).fill = STATUS_FILLS[status]
        for column, tolerance in ((4, data["meta"]["qty_tolerance"]), (7, data["meta"]["amount_tolerance"])):
            value = summary_ws.cell(row_index, column).value
            if isinstance(value, (int, float)) and abs(value) > tolerance:
                summary_ws.cell(row_index, column).fill = DIFF_FILL

    detail_ws = wb.create_sheet("明细差异")
    detail_ws.append(data["detail_headers"])
    for row in data["detail"]:
        detail_ws.append(row)
    style_data_sheet(
        detail_ws,
        [18, 23, 30, 14, 14, 14, 16, 16, 16, 24],
        qty_columns={4, 5, 6},
        amount_columns={7, 8, 9},
    )
    add_table(detail_ws, "ReconciliationDetails")

    unmatched_ws = wb.create_sheet("未匹配文件")
    unmatched_ws.append(data["unmatched_headers"])
    for row in data["unmatched"]:
        unmatched_ws.append(row)
    style_data_sheet(unmatched_ws, [55, 36, 60])
    add_table(unmatched_ws, "UnmatchedStatements")

    status_counts: dict[str, int] = defaultdict(int)
    for row in data["summary"]:
        status_counts[row[8]] += 1
    notes_ws = wb.create_sheet("说明")
    notes_rows = [
        ["项目", "内容"],
        ["总表", data["meta"]["main_path"]],
        ["子表目录", data["meta"]["statement_dir"]],
        ["核对月份", data["meta"]["month"] or "未指定（优先可见且非模板工作表）"],
        ["总表供应商数", data["meta"]["main_supplier_count"]],
        ["子表文件数", data["meta"]["statement_file_count"]],
        ["完成", status_counts.get("完成", 0)],
        ["金额/数量不一致", status_counts.get("未完成-金额/数量不一致", 0)],
        ["未找到对账单", status_counts.get("未完成-未找到对账单", 0)],
        ["需复核", status_counts.get("需复核", 0)],
        ["完成规则", f"数量差异绝对值 <= {data['meta']['qty_tolerance']}，金额差异绝对值 <= {data['meta']['amount_tolerance']}"],
        ["编码规则", "编码差异写入明细；若供应商总数量和总金额一致，汇总仍可判为完成。"],
        ["工作表选择", "优先读取可见、指定月份且名称不含‘格式/模板’的工作表，避免把隐藏模板或历史月份重复计入。"],
    ]
    for row in notes_rows:
        notes_ws.append(row)
    style_data_sheet(notes_ws, [23, 95])
    add_table(notes_ws, "ReconciliationNotes")
    notes_ws.page_setup.orientation = "portrait"

    wb.properties.title = "供应商对账核对结果"
    wb.properties.subject = "总表与供应商子表数量、金额及编码差异"
    wb.save(output_path)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="读取采购总表，递归匹配供应商对账单，并输出数量、金额和编码差异。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "示例：\n"
            "  python3 reconcile_statements.py --main 总表.xlsx --statements 对账单目录 "
            "--output 核对结果.xlsx --month 2026-6"
        ),
    )
    parser.add_argument("--main", type=Path, help="采购总表路径；不传时自动查找当前脚本目录的 采购明细表*.xlsx")
    parser.add_argument("--statements", type=Path, default=DEFAULT_STATEMENT_DIR, help="供应商对账单目录，会递归搜索子目录")
    parser.add_argument("--output", type=Path, default=DEFAULT_RESULT_XLSX, help="输出 Excel 路径")
    parser.add_argument("--month", help="核对月份，例如 2026-6；用于选择正确工作表")
    parser.add_argument("--main-sheet", help="总表工作表名；不传时使用活动工作表")
    parser.add_argument("--main-header-row", type=int, default=3, help="总表表头行，默认 3")
    parser.add_argument("--main-data-row", type=int, default=4, help="总表数据起始行，默认 4")
    parser.add_argument("--qty-tolerance", type=float, default=QTY_TOLERANCE, help="数量差异容差")
    parser.add_argument("--amount-tolerance", type=float, default=AMOUNT_TOLERANCE, help="金额差异容差，默认 0.01")
    parser.add_argument("--aliases-json", type=Path, help="可选供应商别名 JSON")
    parser.add_argument("--json-output", type=Path, default=DEFAULT_DATA_JSON, help="审计用 JSON 输出路径")
    parser.add_argument("--no-json", action="store_true", help="不生成审计 JSON")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    main_path = (args.main or find_default_main()).resolve()
    statement_dir = args.statements.resolve()
    output_path = args.output.resolve()
    aliases = load_aliases(args.aliases_json.resolve() if args.aliases_json else None)
    data = build_output(
        main_path=main_path,
        statement_dir=statement_dir,
        output_xlsx=output_path,
        month=args.month,
        main_sheet=args.main_sheet,
        main_header_row=args.main_header_row,
        main_data_row=args.main_data_row,
        qty_tolerance=args.qty_tolerance,
        amount_tolerance=args.amount_tolerance,
        aliases=aliases,
    )
    write_workbook(data, output_path)
    if not args.no_json:
        json_path = args.json_output.resolve()
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    status_counts = defaultdict(int)
    for row in data["summary"]:
        status_counts[row[8]] += 1
    print(json.dumps({"meta": data["meta"], "status_counts": dict(status_counts), "detail_rows": len(data["detail"])}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
